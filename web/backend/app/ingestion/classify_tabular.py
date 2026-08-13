#!/usr/bin/env python3
"""
Step 3 — Tabular schema classification for the E-Rakshak ingestion pipeline.

Targets CSV/Excel files that remain in uploads/raw/ after Step 2 validation.
Reads ONLY the header row (never the full dataset) and routes by column names:

  A_PARTY / B_PARTY              → uploads/processed/cdr/
  Msisdn / Ip Address            → uploads/processed/ipdr/
  UPI / Description / balance_closing → uploads/processed/bank/
  no known schema                → uploads/quarantine/

Non-tabular files (PDF, Word, images) are left in raw/ for Step 4.

Usage:
    python ingestion_p/classify_tabular.py
    python ingestion_p/classify_tabular.py --dry-run
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import shutil
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path

try:
    from setup_staging import UPLOADS_ROOT, create_staging_layout
except ImportError:  # pragma: no cover
    from ingestion_p.setup_staging import UPLOADS_ROOT, create_staging_layout

logger = logging.getLogger(__name__)

TABULAR_EXTENSIONS = {".csv", ".tsv", ".xls", ".xlsx", ".xlsm"}
HEADER_SCAN_ROWS = 30

# Destination category under processed/ (or "quarantine")
CATEGORY_CDR = "cdr"
CATEGORY_IPDR = "ipdr"
CATEGORY_BANK = "bank"
CATEGORY_QUARANTINE = "quarantine"

# Dynamic / extensible schema signals: (pattern, weight, display_label)
# Patterns are matched against normalized headers (substring or regex).
SCHEMA_SIGNALS: dict[str, list[tuple[str, int, str]]] = {
    CATEGORY_CDR: [
        (r"\ba\s*party\b", 6, "A_PARTY"),
        (r"\bb\s*party\b", 6, "B_PARTY"),
        (r"call\s*type", 3, "CALL_TYPE"),
        (r"call\s*date", 3, "CALL_DATE"),
        (r"call\s*duration|^\s*duration\s*$", 2, "DURATION"),
        (r"first\s*cell|imei", 2, "CELL/IMEI"),
        (r"target.*party|calling\s*party", 4, "A_PARTY"),
    ],
    CATEGORY_IPDR: [
        (r"msisdn|landline.*internet", 6, "Msisdn"),
        (r"public\s*ip|source\s*ip|destination\s*ip|ip\s*address", 5, "Ip Address"),
        (r"data\s*volume", 4, "Data Volume"),
        (r"pgw|session\s*duration|access\s*point", 3, "IPDR session"),
    ],
    CATEGORY_BANK: [
        (r"\bupi\b", 5, "UPI"),
        (r"description|narration|particulars", 4, "Description"),
        (r"balance\s*closing|closing\s*balance", 5, "balance_closing"),
        (r"\bbalance\b", 2, "Balance"),
        (r"debit|credit|withdrawal|deposit|\bdr\s*amt\b|\bcr\s*amt\b|\bdr\b|\bcr\b", 3, "Debit/Credit"),
        (r"transaction\s*(id|date|amount)|tran\s*(id|date|type)|account\s*number|\bac\s*no\b|\bac\s*name\b", 3, "Txn/Account"),
        (r"\bifsc\b|value\s*date|value\s*dt|pstd\s*dt", 2, "IFSC/ValueDate"),
        (r"\bcust\s*id\b|\bcrncy\b", 2, "Core banking"),
    ],
}



@dataclass
class ClassificationResult:
    path: str
    original_name: str
    category: str
    headers: list[str]
    matched_signals: list[str]
    status: str  # "classified" | "quarantined" | "skipped_nontabular" | "failed"
    reason: str
    destination: str | None = None


@dataclass
class ClassificationReport:
    scanned_at: str
    raw_dir: str
    processed_dir: str
    quarantine_dir: str
    total_seen: int = 0
    tabular: int = 0
    classified: int = 0
    quarantined: int = 0
    skipped_nontabular: int = 0
    failed: int = 0
    results: list[ClassificationResult] = field(default_factory=list)


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\ufeff", "").strip().lower()
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip(" :.-")


def read_csv_header_only(path: Path) -> list[str]:
    """
    Find the best header row in the first HEADER_SCAN_ROWS lines.
    Does not load the full file into memory.
    """
    encodings = ("utf-8-sig", "utf-8", "latin-1", "cp1252")
    last_err: Exception | None = None
    cue_words = (
        "party", "date", "time", "duration", "imei", "msisdn", "upi",
        "debit", "credit", "narration", "description", "balance", "ip",
        "account", "call", "cell",
    )
    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as fh:
                sample = fh.read(65536)
                fh.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(fh, dialect)
                best_row: list[str] = []
                best_score = -1
                for idx, row in enumerate(reader):
                    if idx >= HEADER_SCAN_ROWS:
                        break
                    cells = [normalize_header(c) for c in row if str(c).strip() != ""]
                    if len(cells) < 2:
                        continue
                    score = sum(1 for c in cells for w in cue_words if w in c)
                    # Prefer non-numeric-looking header rows
                    numeric = sum(
                        1
                        for c in cells
                        if re.fullmatch(r"[\d.,]+", c.replace(" ", ""))
                    )
                    if numeric / len(cells) > 0.5:
                        continue
                    if score > best_score:
                        best_score = score
                        best_row = cells
                return best_row
        except UnicodeDecodeError as exc:
            last_err = exc
            continue
    if last_err:
        raise last_err
    return []


def read_excel_header_only(path: Path) -> list[str]:
    """Read-only scan of early rows to locate header (not full workbook load)."""
    from openpyxl import load_workbook

    cue_words = (
        "party", "date", "time", "duration", "imei", "msisdn", "upi",
        "debit", "credit", "narration", "description", "balance", "ip",
        "account", "call", "cell",
    )
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        best_row: list[str] = []
        best_score = -1
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= HEADER_SCAN_ROWS:
                break
            cells = [
                normalize_header(c)
                for c in row
                if c is not None and str(c).strip() not in ("", "nan")
            ]
            if len(cells) < 2:
                continue
            score = sum(1 for c in cells for w in cue_words if w in c)
            if score > best_score:
                best_score = score
                best_row = cells
        return best_row
    finally:
        wb.close()


def read_xls_header_only(path: Path) -> list[str]:
    """
    Legacy .xls: use pandas with nrows=0 so only the header is parsed.
    Still avoids loading the data body.
    """
    import pandas as pd

    df = pd.read_excel(path, nrows=0, engine="xlrd")
    return [normalize_header(c) for c in df.columns]


def extract_headers(path: Path) -> list[str]:
    """Return normalized header names from the first row only."""
    ext = path.suffix.lower()
    if ext in {".csv", ".tsv"}:
        return read_csv_header_only(path)
    if ext == ".xls":
        return read_xls_header_only(path)
    if ext in {".xlsx", ".xlsm"}:
        return read_excel_header_only(path)
    raise ValueError(f"not a tabular extension: {ext}")


def classify_from_headers(headers: list[str]) -> tuple[str, list[str], str]:
    """
    Score-based schema classification (extensible via SCHEMA_SIGNALS).
    Returns (category, matched_signals, reason).
    """
    joined = " | ".join(headers)
    scores: dict[str, int] = {cat: 0 for cat in SCHEMA_SIGNALS}
    labels: dict[str, list[str]] = {cat: [] for cat in SCHEMA_SIGNALS}

    for category, rules in SCHEMA_SIGNALS.items():
        seen_labels: set[str] = set()
        for pattern, weight, label in rules:
            if re.search(pattern, joined, flags=re.IGNORECASE):
                scores[category] += weight
                if label not in seen_labels:
                    labels[category].append(label)
                    seen_labels.add(label)

    ranked = sorted(scores.items(), key=lambda kv: kv[1], reverse=True)
    best, best_score = ranked[0]
    second = ranked[1][1] if len(ranked) > 1 else 0

    if best_score < 5:
        return CATEGORY_QUARANTINE, [], f"no schema matched strongly (scores={scores})"

    if best_score - second < 2 and second >= 5:
        # Near-tie: prefer more specific telecom markers in fixed priority
        for pref in (CATEGORY_CDR, CATEGORY_IPDR, CATEGORY_BANK):
            if scores[pref] >= second:
                return pref, labels[pref], f"near-tie resolved to {pref}; scores={scores}"
        return CATEGORY_QUARANTINE, [], f"ambiguous scores={scores}"

    return best, labels[best], f"matched {best} (score={best_score})"



def _unique_destination(dest_dir: Path, filename: str) -> Path:
    candidate = dest_dir / filename
    if not candidate.exists():
        return candidate
    stem = Path(filename).stem
    suffix = Path(filename).suffix
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%f")
    return dest_dir / f"{stem}__{stamp}{suffix}"


def _move_with_sidecar(
    src: Path,
    dest: Path,
    *,
    category: str,
    headers: list[str],
    matched_signals: list[str],
    reason: str,
) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(src), str(dest))
    meta = dest.with_suffix(dest.suffix + ".classify.json")
    meta.write_text(
        json.dumps(
            {
                "original_name": src.name,
                "category": category,
                "headers": headers,
                "matched_signals": matched_signals,
                "reason": reason,
                "classified_at": datetime.now(timezone.utc).isoformat(),
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def classify_tabular_file(path: Path) -> ClassificationResult:
    """Classify one tabular file from its first-row headers only."""
    name = path.name
    try:
        headers = extract_headers(path)
        if not headers:
            return ClassificationResult(
                path=str(path),
                original_name=name,
                category=CATEGORY_QUARANTINE,
                headers=[],
                matched_signals=[],
                status="failed",
                reason="empty or missing header row",
            )
        category, signals, reason = classify_from_headers(headers)
        status = "classified" if category != CATEGORY_QUARANTINE else "quarantined"
        return ClassificationResult(
            path=str(path),
            original_name=name,
            category=category,
            headers=headers,
            matched_signals=signals,
            status=status,
            reason=reason,
        )
    except Exception as exc:  # noqa: BLE001
        return ClassificationResult(
            path=str(path),
            original_name=name,
            category=CATEGORY_QUARANTINE,
            headers=[],
            matched_signals=[],
            status="failed",
            reason=f"{type(exc).__name__}: {exc}",
        )


def classify_raw_tabular(
    raw_dir: Path | None = None,
    processed_dir: Path | None = None,
    quarantine_dir: Path | None = None,
    *,
    move_files: bool = True,
) -> ClassificationReport:
    """
    Scan uploads/raw/ for validated CSV/Excel files, classify by headers only,
    and route into processed/cdr|ipdr|bank or quarantine.
    """
    create_staging_layout()

    raw = Path(raw_dir) if raw_dir else UPLOADS_ROOT / "raw"
    processed = Path(processed_dir) if processed_dir else UPLOADS_ROOT / "processed"
    quarantine = Path(quarantine_dir) if quarantine_dir else UPLOADS_ROOT / "quarantine"

    for d in (raw, processed, quarantine):
        d.mkdir(parents=True, exist_ok=True)
    for sub in (CATEGORY_CDR, CATEGORY_IPDR, CATEGORY_BANK):
        (processed / sub).mkdir(parents=True, exist_ok=True)

    report = ClassificationReport(
        scanned_at=datetime.now(timezone.utc).isoformat(),
        raw_dir=str(raw),
        processed_dir=str(processed),
        quarantine_dir=str(quarantine),
    )

    files = sorted(p for p in raw.iterdir() if p.is_file() and p.name != ".gitkeep")
    report.total_seen = len(files)

    for path in files:
        if path.suffix.lower() not in TABULAR_EXTENSIONS:
            report.skipped_nontabular += 1
            report.results.append(
                ClassificationResult(
                    path=str(path),
                    original_name=path.name,
                    category="",
                    headers=[],
                    matched_signals=[],
                    status="skipped_nontabular",
                    reason="left in raw/ for unstructured classification (Step 4)",
                )
            )
            continue

        report.tabular += 1
        result = classify_tabular_file(path)

        if result.category == CATEGORY_QUARANTINE or result.status == "failed":
            report.quarantined += 1
            dest_dir = quarantine
            if result.status != "failed":
                result.status = "quarantined"
        else:
            report.classified += 1
            dest_dir = processed / result.category

        dest = _unique_destination(dest_dir, path.name)
        if move_files:
            _move_with_sidecar(
                path,
                dest,
                category=result.category or CATEGORY_QUARANTINE,
                headers=result.headers,
                matched_signals=result.matched_signals,
                reason=result.reason,
            )
            result.destination = str(dest)
        else:
            result.destination = str(dest)

        report.results.append(result)

    return report


def save_report(report: ClassificationReport, path: Path | None = None) -> Path:
    out = path or (UPLOADS_ROOT / "classification_report.json")
    payload = {
        **{k: v for k, v in asdict(report).items() if k != "results"},
        "results": [asdict(r) for r in report.results],
    }
    out.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return out


def _print_report(report: ClassificationReport) -> None:
    print("=== Step 3: Tabular Schema Classification ===")
    print(f"raw         : {report.raw_dir}")
    print(f"processed   : {report.processed_dir}")
    print(f"quarantine  : {report.quarantine_dir}")
    print(f"seen        : {report.total_seen}")
    print(f"tabular     : {report.tabular}")
    print(f"classified  : {report.classified}")
    print(f"quarantined : {report.quarantined}")
    print(f"nontabular  : {report.skipped_nontabular} (left in raw/)")
    print(f"failed      : {report.failed}")
    print()
    for r in report.results:
        if r.status == "skipped_nontabular":
            print(f"  [SKIP] {r.original_name}")
            continue
        label = r.category.upper() if r.category else "UNKNOWN"
        print(f"  [{label:12}] {r.original_name}")
        print(f"               signals={r.matched_signals}  reason={r.reason}")
        if r.destination:
            print(f"               -> {r.destination}")


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    parser = argparse.ArgumentParser(
        description="Classify validated CSV/Excel files by header row only"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Classify and report only; do not move files",
    )
    args = parser.parse_args()
    report = classify_raw_tabular(move_files=not args.dry_run)
    _print_report(report)
    out = save_report(report)
    print(f"\nReport saved: {out}")
    if args.dry_run:
        print("(dry-run: no files were moved)")


if __name__ == "__main__":
    main()
