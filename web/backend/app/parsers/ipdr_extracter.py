#!/usr/bin/env python3
"""
Simple IPDR/CDR unified extractor -- NO LLM, exact-header mapping only.

Usage:
    python3 ipdr_extracter.py /path/to/input_folder  [output.xlsx]

What it does (in plain terms)
------------------------------
1. Walks a folder of .xlsx/.xls/.csv/.pdf/.docx files.
2. For every table it finds, looks at the header row and maps each column
   to a fixed, known field name using a simple lookup table (HEADER_MAP
   below) built from the real IPDR/CDR header formats you gave me. No
   fuzzy scoring, no regex guessing, no content-sniffing, no LLM -- if a
   header isn't in the list, its column is kept as-is under its own name
   (nothing is dropped) but isn't force-mapped to a canonical field.
3. Every row becomes one row in a single output sheet, with lineage
   (source file / sheet / row) and a dedup hash (drops exact repeats of
   the same event across files, e.g. an IPDR grid pasted into both an
   Excel export and an email PDF).

That's the whole pipeline. Add new header formats by adding lines to
HEADER_MAP -- nothing else needs to change.
"""

import os
import re
import sys
import hashlib
import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import docx as python_docx
except ImportError:
    python_docx = None


# ============================================================================
# 1. HEADER -> CANONICAL FIELD  (exact lookup, built from your real formats)
# ============================================================================
# Key = header text, lower-cased and whitespace-collapsed. Add more lines
# here whenever you meet a new file format -- that's the only edit needed.

def _clean_msisdn(v):
    """Keep only real-looking Indian mobile numbers: exactly 10 digits,
    starting 6-9, after stripping spaces/dashes/+91/91 prefixes. Anything
    else (ranks, counts, stray small numbers) is dropped, not guessed at."""
    if v is None:
        return None
    digits = re.sub(r"\D", "", str(v))
    if digits.startswith("91") and len(digits) == 12:
        digits = digits[2:]
    return digits if re.fullmatch(r"[6-9]\d{9}", digits) else None


def _clean_imei(v):
    if v is None:
        return None
    digits = re.sub(r"\D", "", str(v))
    return digits if re.fullmatch(r"\d{15}", digits) else None


def _clean_imsi(v):
    if v is None:
        return None
    digits = re.sub(r"\D", "", str(v))
    return digits if re.fullmatch(r"\d{14,15}", digits) else None


def _duration_to_seconds(v):
    """Accept either a plain number of seconds or an 'HH:MM:SS'/'MM:SS'
    text value and normalize both to an integer number of seconds."""
    if v is None or str(v).strip() == "":
        return None
    s = str(v).strip()
    if re.fullmatch(r"\d+", s):
        return int(s)
    if re.fullmatch(r"\d{1,2}:\d{2}:\d{2}", s):
        h, m, sec = (int(x) for x in s.split(":"))
        return h * 3600 + m * 60 + sec
    if re.fullmatch(r"\d{1,2}:\d{2}", s):
        m, sec = (int(x) for x in s.split(":"))
        return m * 60 + sec
    return None  # unrecognized shape -> drop rather than store text in a numeric field


def norm(h):
    return re.sub(r"\s+", " ", str(h).strip().lower()).strip(" :.-")


HEADER_MAP = {
    # ---- format: CDR summary/manifest ----
    "cdr name": "cdr_name",
    "cdr type": "cdr_type",
    "network type": "network_type",
    "circle": "circle",
    "operator": "operator",
    "records": "records",
    "remark": "remark",

    # ---- format: subscriber/KYC style (No. / IP Address / username ...) ----
    "no.": "sr_no",
    "no": "sr_no",
    "ip address": "ip_address",
    "date": "date",
    "time(ist)": "time",
    "username": "username",
    "firstname": "firstname",
    "fulladdress": "full_address",
    "mobile": "msisdn",

    # ---- format: IP / VALUE / F DATE / F TIME / T DATE / T TIME ----
    "ip": "ip_type",
    "value": "ip_address",
    "f date": "start_date",
    "f time": "start_time",
    "t date": "end_date",
    "t time": "end_time",
    "l date": "end_date",
    "l time": "end_time",
    "start time(ist)": "start_time",
    "end time(ist)": "end_time",
    "ip (jio)": "source_ip",
    "fir/case no": "case_no",
    "fir case no": "case_no",
    "case no": "case_no",
    "case no.": "case_no",

    # ---- format: full IPDR (NAT / port / session) ----
    "landline/msisdn/mdn/leased circuit id for internet access": "msisdn",
    "source ip address (ip address assigned/translated)": "source_ip",
    "source port (port assigned/translated)": "source_port",
    "public ip address": "public_ip",
    "public ip port": "public_ip_port",
    "destination ip address": "destination_ip",
    "destination port": "destination_port",
    "start date of public ip address allocation (dd/mm/yyyy)": "start_date",
    "ist start time of public ip address allocation (hh:mm:ss)": "start_time",
    "end date of public ip address allocation (dd/mm/yyyy)": "end_date",
    "ist end time of public ip address allocation (hh:mm:ss)": "end_time",
    "static/dynamic ip address allocation": "ip_allocation_type",
    "user id for internet access based on authentication": "user_id",
    "source mac-id address/other device identification number": "mac_id",
    "imsi": "imsi",
    "pgw ip address": "pgw_ip",
    "access point name": "apn",
    "apn": "apn",
    "cgi id": "cgi_id",
    "time1 (dd/mm/yyyy hh:mm:ss)": "event_time",
    "roaming circle indicator": "roaming_indicator",
    "roaming circle": "roaming_circle",
    "session duration": "duration_sec",
    "data volume up link": "data_volume_up",
    "data volume down link": "data_volume_down",

    # ---- format: CDR (voice call detail) ----
    "a party": "msisdn_a",
    "b party": "msisdn_b",
    "a party no.": "msisdn_a",
    "b party no.": "msisdn_b",
    "time": "time",
    "duration": "duration_sec",
    "call type": "call_type",
    "first cell id a": "first_cell_id",
    "last cell id a": "last_cell_id",
    "imei a": "imei",
    "imsi a": "imsi",
    "first cell id a address": "cell_address",
    "ip details": "ip_address",
    "roaming a": "roaming_indicator",
    "latitude": "latitude",
    "longitude": "longitude",

    # ---- format: Reports sheet (Max Call / Max Duration / Max Location / Max IMEI) ----
    "number": "msisdn",
    "description": "subscriber_description",
    "type": "subscriber_type",
    "total": "total_records",
    "internet": "internet_records",
    "start date": "start_date",
    "start time": "start_time",
    "end date": "end_date",
    "end time": "end_time",
    "spam": "spam_flag",
    "total duration": "total_duration",
    # Note: 'A PARTY'/'B PARTY' headers in voice CDR sheets = phone numbers → msisdn_a/msisdn_b
    # In Max Duration report they = call counts but we cannot disambiguate at header level.
    # They map to msisdn_a/msisdn_b (voice CDR meaning) which is defined earlier above.
    "a + b": "call_direction",
    "tower number": "cell_id",
    "tower address": "cell_address",
    "imei number": "imei",
    "handset details": "handset_details",
    "used numbers": "linked_numbers_count",
    "numbers": "linked_numbers",
    "used imsi": "linked_imsi_count",

    # ---- additive aliases (Step 1 additions — do not remove existing lines) ----
    "device details": "handset_details",
    "subscriber type": "subscriber_type",
    "spam flag": "spam_flag",
    "linked numbers": "linked_numbers",
    "linked numbers count": "linked_numbers_count",
    "linked imsi count": "linked_imsi_count",

    # ---- common short aliases ----
    "msisdn": "msisdn",
    "msisdn/ani": "msisdn",
    "imei": "imei",
    "ip address": "ip_address",
}

# The output is split into two sheets for clean presentation.
PRIMARY_COLUMNS = [
    "record_id",
    "msisdn",
    "ip_address",
    "cell_id",
    "start_date", "start_time", "end_date", "end_time", "duration_sec",
    "data_volume_up", "data_volume_down",
]

SECONDARY_COLUMNS = [
    "record_id",
    "name", "msisdn_b", "imei", "imsi",
    "full_address",
    "destination_ip", "destination_port",
    "latitude", "longitude",
    "roaming_circle",
    "event_hash", "dup_status",
    "source_file", "source_sheet_or_page", "source_row_ref",
]

FONT = "Arial"


# ============================================================================
# 2. FILE PARSERS  -- read every table, map headers via HEADER_MAP
# ============================================================================

def clean_val(v):
    """Collapse newlines/tabs/extra spaces in a cell value (common in PDF extraction)."""
    if v is None:
        return v
    s = str(v)
    s = re.sub(r"[\r\n\t]+", " ", s)   # newlines → space
    s = re.sub(r" {2,}", " ", s)        # multiple spaces → one
    return s.strip()


def build_row(headers, values, source_file, source_loc, row_ref):
    """headers/values are parallel lists for one row. Map every recognized
    header to its canonical field; keep unrecognized columns too (under
    their original header) so nothing is silently lost."""
    row = {}
    for h, v in zip(headers, values):
        if v is None or str(v).strip() == "" or str(v).strip().lower() == "nan":
            continue
        if h is None or str(h).strip() == "":
            continue
        v = clean_val(v)          # strip embedded newlines from PDF text
        if not v or str(v).strip().lower() == "nan":
            continue
        field = HEADER_MAP.get(norm(h))
        if field:
            row[field] = v
        else:
            row[f"extra::{str(h).strip()}"] = v
    if not row:
        return None
    row["source_file"] = source_file
    row["source_sheet_or_page"] = source_loc
    row["source_row_ref"] = row_ref
    return row


def find_header_row(rows, max_scan=30):
    """Return the index of the row with the most HEADER_MAP hits (min 2)."""
    best_idx, best_hits = None, 1  # need at least 2 hits to count as a header
    for i, r in enumerate(rows[:max_scan]):
        hits = sum(1 for c in r if c is not None and norm(c) in HEADER_MAP)
        if hits > best_hits:
            best_idx, best_hits = i, hits
    return best_idx


def _read_sheet_with_merges_filled(ws):
    """Read an openpyxl worksheet into a DataFrame, filling every cell that
    is part of a merged range with that range's top-left value. Cells that
    are NOT part of any merge are left exactly as they are (including
    genuinely blank ones) -- this only restores data Excel is displaying
    as 'shared downward', it never invents data that wasn't there."""
    data = [[c.value for c in row] for row in ws.iter_rows()]
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        top_left_value = data[min_row - 1][min_col - 1]
        for r in range(min_row - 1, max_row):
            for c in range(min_col - 1, max_col):
                data[r][c] = top_left_value
    return pd.DataFrame(data, dtype=object)


def parse_excel_or_csv(path, log):
    out_rows = []
    ext = os.path.splitext(path)[1].lower()
    try:
        sheets = {}
        if ext == ".csv":
            sheets["csv"] = pd.read_csv(path, header=None, dtype=object, engine="python", on_bad_lines="skip")
        else:
            # Use openpyxl directly so merged-cell ranges are visible.
            # _read_sheet_with_merges_filled() fills every non-top-left cell
            # in a merge with the top-left value (e.g. one IP shared across
            # several subscriber rows) before handing a normal DataFrame to
            # the rest of the pipeline. Unmerged blank cells are unchanged.
            wb_raw = openpyxl.load_workbook(path, data_only=True)
            for s in wb_raw.sheetnames:
                sheets[s] = _read_sheet_with_merges_filled(wb_raw[s])
    except Exception as e:
        log.append((os.path.basename(path), f"FAILED: {e}"))
        return out_rows

    _SKIP_SHEET_PATTERNS = ("max ", "top ", "summary", "rank", "stat", "leaderboard")
    n_rows = 0
    for sheet_name, df in sheets.items():
        if any(p in sheet_name.strip().lower() for p in _SKIP_SHEET_PATTERNS):
            log.append((os.path.basename(path), f"SKIPPED sheet '{sheet_name}' - looks like a summary/leaderboard sheet, not raw records"))
            continue
        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
        if df.empty:
            continue
        rows = df.values.tolist()
        hdr_idx = find_header_row(rows)
        if hdr_idx is None:
            continue  # no recognizable IPDR/CDR header in this sheet -> skip (kept simple, on purpose)
        headers = rows[hdr_idx]
        for i, r in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
            if all(c is None or str(c).strip() == "" for c in r):
                continue
            row = build_row(headers, r, os.path.basename(path), sheet_name, f"row {i}")
            if row:
                out_rows.append(row)
                n_rows += 1
    log.append((os.path.basename(path), f"OK - {n_rows} row(s)"))
    return out_rows


def parse_pdf(path, log):
    out_rows = []
    if pdfplumber is None:
        log.append((os.path.basename(path), "FAILED: pdfplumber not available"))
        return out_rows
    n_rows = 0
    try:
        with pdfplumber.open(path) as pdf:
            for pno, page in enumerate(pdf.pages, start=1):
                for tno, table in enumerate(page.extract_tables() or [], start=1):
                    if not table or len(table) < 2 or len(table[0]) < 2:
                        continue
                    hdr_idx = find_header_row(table)
                    if hdr_idx is None:
                        continue
                    headers = table[hdr_idx]
                    for i, r in enumerate(table[hdr_idx + 1:], start=hdr_idx + 2):
                        if all(c is None or str(c).strip() == "" for c in r):
                            continue
                        row = build_row(headers, r, os.path.basename(path), f"page {pno}", f"row {i}")
                        if row:
                            out_rows.append(row)
                            n_rows += 1
        log.append((os.path.basename(path), f"OK - {n_rows} row(s)"))
    except Exception as e:
        log.append((os.path.basename(path), f"FAILED: {e}"))
    return out_rows


def parse_docx(path, log):
    out_rows = []
    if python_docx is None:
        log.append((os.path.basename(path), "FAILED: python-docx not available"))
        return out_rows
    n_rows = 0
    try:
        doc = python_docx.Document(path)
        for tno, table in enumerate(doc.tables, start=1):
            rows = [[c.text for c in r.cells] for r in table.rows]
            if len(rows) < 2:
                continue
            hdr_idx = find_header_row(rows)
            if hdr_idx is None:
                continue
            headers = rows[hdr_idx]
            for i, r in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
                if not any(c.strip() for c in r if c):
                    continue
                row = build_row(headers, r, os.path.basename(path), f"table {tno}", f"row {i}")
                if row:
                    out_rows.append(row)
                    n_rows += 1
        log.append((os.path.basename(path), f"OK - {n_rows} row(s)"))
    except Exception as e:
        log.append((os.path.basename(path), f"FAILED: {e}"))
    return out_rows


# ============================================================================
# 3. DATE/TIME NORMALIZATION + DEDUP HASH
# ============================================================================

def to_datetime(date_val, time_val):
    if date_val is None:
        return None

    # BUG FIX 1: If date_val is already a datetime/Timestamp, extract directly
    if isinstance(date_val, (datetime.datetime, pd.Timestamp)):
        base = pd.Timestamp(date_val)
        y, mo, da = base.year, base.month, base.day
    else:
        ds = str(date_val).strip()
        try:
            if re.fullmatch(r"\d{8}", ds):          # YYYYMMDD
                y, mo, da = int(ds[:4]), int(ds[4:6]), int(ds[6:8])
            else:                                     # dd/mm/yyyy or similar
                ds_clean = re.sub(r"\s*\(.*?\)$", "", ds).strip()
                # Try ISO format first (avoids pandas dayfirst warning for YYYY-MM-DD strings)
                dt = pd.to_datetime(ds_clean, format="%Y-%m-%d %H:%M:%S", errors="coerce")
                if pd.isna(dt):
                    dt = pd.to_datetime(ds_clean, format="%Y-%m-%d", errors="coerce")
                if pd.isna(dt):
                    dt = pd.to_datetime(ds_clean, dayfirst=True, errors="coerce")
                if pd.isna(dt):
                    return None
                y, mo, da = dt.year, dt.month, dt.day
        except Exception:
            return None

    # BUG FIX 2: Excel stores times as a float fraction of a day (e.g. 0.5 = noon).
    # Detect this and convert properly instead of stripping digits from the float string.
    try:
        if time_val is None:
            hh, mm, ss = 0, 0, 0
        elif isinstance(time_val, (int, float)) and not isinstance(time_val, bool) and 0.0 <= float(time_val) < 1.0:
            total_sec = round(float(time_val) * 86400)
            hh = total_sec // 3600
            mm = (total_sec % 3600) // 60
            ss = total_sec % 60
        elif isinstance(time_val, datetime.time):
            hh, mm, ss = time_val.hour, time_val.minute, time_val.second
        elif isinstance(time_val, (datetime.datetime, pd.Timestamp)):
            t = pd.Timestamp(time_val)
            hh, mm, ss = t.hour, t.minute, t.second
        else:
            ts = re.sub(r"\D", "", str(time_val).strip())
            ts = ts.zfill(6)[:6]
            hh, mm, ss = int(ts[0:2]), int(ts[2:4]), int(ts[4:6])
        return datetime.datetime(y, mo, da, hh, mm, ss)
    except Exception:
        return None


def normalize_msisdn(v):
    """
    Normalize a phone number to a clean 10-digit Indian mobile number.
    - Strips country code 91 / +91 from 12-digit numbers starting with 91.
    - Strips non-digit characters (spaces, dashes, +).
    - Returns None if result is not 10 digits (not a valid mobile number).
    """
    if not v:
        return v
    s = re.sub(r"[^\d]", "", str(v).strip())   # keep digits only
    # Strip country code: 91XXXXXXXXXX (12 digits starting with 91)
    if len(s) == 12 and s.startswith("91"):
        s = s[2:]
    # Some files prefix with 0 (STD code), strip leading 0
    if len(s) == 11 and s.startswith("0"):
        s = s[1:]
    # Return only if it looks like a valid 10-digit Indian mobile
    if len(s) == 10 and s[0] in "6789":
        return s
    # Not a phone number (e.g. "0", "13" call-counts) — return original
    return v


def normalize_date(v):
    """
    Normalize any date string to YYYY-MM-DD.
    Handles: dd/mm/yyyy, YYYY-MM-DD, YYYY-MM-DD HH:MM:SS, timestamps.
    """
    if not v:
        return v
    s = str(v).strip()
    # Already YYYY-MM-DD (possibly with time part appended)
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    # dd/mm/yyyy  or  dd-mm-yyyy
    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})", s)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    # YYYYMMDD
    if re.fullmatch(r"\d{8}", s):
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    return v   # return as-is if we can't parse


def normalize_imsi(v):
    """Strip count suffixes like '405870182365083 (13)' → '405870182365083'."""
    if not v:
        return v
    return re.sub(r"\s*\(\d+\)\s*$", "", str(v).strip())


def normalize_imei(v):
    """IMEI should be 15 digits. Some sources give 14 (drop check digit). Keep as-is but clean."""
    if not v:
        return v
    s = re.sub(r"[^\d]", "", str(v).strip())
    return s if len(s) in (14, 15) else v


def normalize_row(row):
    """
    Merge alias fields + normalize ALL values to consistent formats:
      - Phone numbers  → 10-digit (strip +91 country code)
      - Dates          → YYYY-MM-DD
      - IMSI           → strip trailing count suffix
      - IMEI           → digits only
      - ip_address ↔ source_ip  (cross-fill aliases)
      - date/time ↔ start_date/start_time  (cross-fill aliases)
      - user_id → msisdn  (IPDR user_id is often MSISDN with country code)
    """
    def fill(dst, src):
        if row.get(dst) in (None, "") and row.get(src) not in (None, ""):
            row[dst] = row[src]

    # ── 1. Phone number normalization ──────────────────────────────────────
    for col in ("msisdn", "msisdn_a", "msisdn_b", "user_id"):
        if row.get(col):
            row[col] = normalize_msisdn(row[col])

    # ── 2. Date normalization (all raw date fields before start/end merge) ──
    for col in ("date", "start_date", "end_date"):
        if row.get(col):
            row[col] = normalize_date(row[col])

    # ── 3. IMSI cleanup ─────────────────────────────────────────────────────
    if row.get("imsi"):
        row["imsi"] = normalize_imsi(row["imsi"])

    # ── 4. IMEI cleanup ─────────────────────────────────────────────────────
    if row.get("imei"):
        row["imei"] = normalize_imei(row["imei"])

    # ── 5. IP address aliases (cross-fill) ──────────────────────────────────
    fill("source_ip",  "ip_address")
    fill("ip_address", "source_ip")

    # ── 6. Date/Time aliases (cross-fill to unify bare Date/Time columns) ───
    fill("start_date", "date")
    fill("start_time", "time")
    fill("date",       "start_date")
    fill("time",       "start_time")

    # ── 7. MSISDN aliases ───────────────────────────────────────────────────
    # user_id in full IPDR = MSISDN with country code — promote after normalizing
    fill("msisdn", "user_id")

    return row



def finalize_row(row, counter):
    """Fill start/end datetime from whichever date+time fields are present,
    consolidate alias fields, compute duration if missing, and a dedup hash."""
    counter[0] += 1
    row["record_id"] = f"REC-{counter[0]:05d}"

    # --- merge redundant name columns into a single 'name' field ---
    # Prefer firstname (real name) over username (login handle); use username only as fallback.
    if not row.get("name"):
        row["name"] = row.get("firstname") or row.get("username") or None

    # --- merge first/last/cgi cell id into a single 'cell_id' field ---
    if not row.get("cell_id"):
        first_c = row.get("first_cell_id")
        last_c  = row.get("last_cell_id")
        cgi     = row.get("cgi_id")
        if first_c and last_c and str(first_c) != str(last_c):
            row["cell_id"] = f"{first_c}→{last_c}"
        elif first_c or last_c or cgi:
            row["cell_id"] = first_c or last_c or cgi

    sdt = to_datetime(row.get("start_date") or row.get("date"), row.get("start_time") or row.get("time"))
    edt = to_datetime(row.get("end_date") or row.get("start_date") or row.get("date"), row.get("end_time"))
    if sdt:
        row["start_date"] = sdt.strftime("%Y-%m-%d")
        row["start_time"] = sdt.strftime("%H:%M:%S")
    if edt:
        row["end_date"] = edt.strftime("%Y-%m-%d")
        row["end_time"] = edt.strftime("%H:%M:%S")

    # Only compute duration if BOTH start AND end times were explicitly in the source data
    # (avoids negative durations when end_time is absent and defaults to 00:00:00)
    if sdt and edt and not row.get("duration_sec"):
        raw_end_time = row.get("end_time")
        # Only calc if the end_time is not midnight default OR dates differ
        had_explicit_end = (raw_end_time and raw_end_time != "00:00:00") or (
            row.get("end_date") and row.get("end_date") != row.get("start_date")
        )
        diff_sec = int((edt - sdt).total_seconds())
        if had_explicit_end and diff_sec >= 0:
            row["duration_sec"] = diff_sec

    # --- fold total_duration into duration_sec if no session duration found ---
    if not row.get("duration_sec") and row.get("total_duration"):
        row["duration_sec"] = row["total_duration"]

    # Normalize all fields (phone numbers, dates, IMEI/IMSI, alias cross-fill)
    # BEFORE hashing so cross-file duplicates are detected correctly
    normalize_row(row)

    # --- strict post-normalization cleanup ---
    for f in ("msisdn", "msisdn_a", "msisdn_b"):
        if row.get(f):
            row[f] = _clean_msisdn(row[f])
    if not row.get("msisdn") and row.get("msisdn_a"):
        row["msisdn"] = row["msisdn_a"]
    if row.get("msisdn_b") and row.get("msisdn_b") == row.get("msisdn"):
        row["msisdn_b"] = None
    if row.get("imei"):
        row["imei"] = _clean_imei(row["imei"])
    if row.get("imsi"):
        row["imsi"] = _clean_imsi(row["imsi"])
    if row.get("duration_sec") is not None:
        row["duration_sec"] = _duration_to_seconds(row["duration_sec"])

    # Hash uses the unified canonical fields (post-consolidation)
    key_fields = ["msisdn", "msisdn_a", "msisdn_b", "imei", "imsi",
                  "source_ip", "public_ip", "destination_ip",
                  "start_date", "start_time", "end_date", "end_time"]
    raw_key = "|".join(str(row.get(k, "")) for k in key_fields)
    row["event_hash"] = hashlib.sha256(raw_key.encode()).hexdigest()[:16]
    return row


# ============================================================================
# 4. MAIN
def run(input_folder, output_path):
    log = []
    all_rows = []
    extra_cols_seen = []

    exts = {".xlsx": "excel", ".xlsm": "excel", ".xls": "excel", ".csv": "csv",
            ".pdf": "pdf", ".docx": "docx"}
    files = []
    for root, _, fnames in os.walk(input_folder):
        for f in fnames:
            ext = os.path.splitext(f)[1].lower()
            if ext in exts:
                files.append(os.path.join(root, f))
    files.sort()
    print(f"Found {len(files)} file(s)")

    for path in files:
        ext = os.path.splitext(path)[1].lower()
        kind = exts[ext]
        print(f"  parsing [{kind}] {os.path.basename(path)}")
        if kind in ("excel", "csv"):
            rows = parse_excel_or_csv(path, log)
        elif kind == "pdf":
            rows = parse_pdf(path, log)
        else:
            rows = parse_docx(path, log)
        all_rows.extend(rows)

    print(f"Total rows extracted: {len(all_rows)}")

    counter = [0]
    for row in all_rows:
        finalize_row(row, counter)
        for k in row:
            if k.startswith("extra::") and k not in extra_cols_seen:
                extra_cols_seen.append(k)

    # dedup (merge any fields the duplicate has but the kept record doesn't,
    # e.g. the FIR/Case No that only appears on the PDF copy of a row)
    seen = {}          # hash -> the kept row dict
    for row in all_rows:
        h = row["event_hash"]
        if h in seen:
            kept_row = seen[h]
            for k, v in row.items():
                if v not in (None, "") and kept_row.get(k) in (None, ""):
                    kept_row[k] = v
            row["dup_status"] = f"DUPLICATE of {kept_row['record_id']}"
        else:
            row["dup_status"] = "UNIQUE"
            seen[h] = row
    kept = [r for r in all_rows if r["dup_status"] == "UNIQUE"]
    dropped = len(all_rows) - len(kept)
    print(f"Unique records kept: {len(kept)}  (duplicates dropped: {dropped})")

    write_excel(output_path, kept, extra_cols_seen, log)
    return output_path


def write_excel(output_path, rows, extra_cols, log):
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    CELL_FONT = Font(name=FONT, size=10)
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def clean_cell(v):
        """Never write the literal string 'nan'/'NaN'/None to a cell — use empty string."""
        if v is None:
            return ""
        if isinstance(v, float) and __import__('math').isnan(v):
            return ""
        if str(v).strip().lower() == "nan":
            return ""
        return v

    def create_data_sheet(wb, title, cols_list):
        ws = wb.create_sheet(title)
        pretty = [c.replace("extra::", "[unmapped] ").replace("_", " ").title() for c in cols_list]
        ws.append(pretty)
        
        for c in range(1, len(cols_list) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"

        for row in rows:
            ws.append([clean_cell(row.get(c, "")) for c in cols_list])
            
        for r in range(2, ws.max_row + 1):
            for c in range(1, len(cols_list) + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = CELL_FONT
                cell.border = BORDER

        for i in range(1, len(cols_list) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

    def create_source_log(wb):
        ws2 = wb.create_sheet("Source_Log")
        ws2.append(["File", "Status"])
        for c in range(1, 3):
            cell = ws2.cell(row=1, column=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
        for f, s in log:
            ws2.append([f, s])
        for r in range(2, ws2.max_row + 1):
            for c in range(1, 3):
                ws2.cell(row=r, column=c).font = CELL_FONT
                ws2.cell(row=r, column=c).border = BORDER
        ws2.column_dimensions["A"].width = 45
        ws2.column_dimensions["B"].width = 90

    # Save Primary Workbook
    wb1 = openpyxl.Workbook()
    wb1.remove(wb1.active)
    create_data_sheet(wb1, "Primary_Info", PRIMARY_COLUMNS)
    create_source_log(wb1)
    wb1.save(output_path)
    print(f"Saved: {output_path}")

    # Save Secondary Workbook
    base, ext = os.path.splitext(output_path)
    sec_path = f"{base}_secondary{ext}"
    wb2 = openpyxl.Workbook()
    wb2.remove(wb2.active)
    create_data_sheet(wb2, "Additional_Info", SECONDARY_COLUMNS + extra_cols)
    create_source_log(wb2)
    wb2.save(sec_path)
    print(f"Saved: {sec_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    in_folder = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.getcwd(), "IPDR_Unified.xlsx")
    if not os.path.isdir(in_folder):
        print(f"ERROR: {in_folder} is not a folder")
        sys.exit(1)
    run(in_folder, out_path)